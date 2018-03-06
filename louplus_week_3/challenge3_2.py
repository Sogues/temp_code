# encoding: utf-8

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def combine():
    wb = load_workbook('courses.xlsx')
    stu = wb['students']
    ti = wb['time']
    ns = wb.create_sheet()
    ns.title = 'combine'
    #ns.append(['创建时间', '课程名称', '学习人数', '学习时间'])
    for s in stu.values:
        for t in ti.values:
            if s[1] == t[1]:
                ns.append(list(s) + [t[2]])
    wb.save('courses.xlsx')


def split():
    wb = load_workbook('courses.xlsx')
    cb = wb['combine']
    ret_dict = {}
    titles=  [item.value for item in cb[1]]
    for c in cb.values:
        if c[0] != '创建时间':
            if c[0].year not in ret_dict.keys():
                n_wb = Workbook()
                ws = n_wb.active
                ws.title = str(c[0].year)
                #ws.append(['创建时间', '课程名称', '学习人数', '学习时间'])
                ws.append(titles)
                ret_dict[c[0].year] = n_wb
            ret_dict[c[0].year].active.append(c)
    for key, wb in ret_dict.items():
        wb.save('{}.xlsx'.format(key))



if __name__ == '__main__':
    combine()
    split()
